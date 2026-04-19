"""
agent5_designer.py: デザイン
"""
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── カラー定義 ──────────────────────────────────────────────────────
COLOR_HEADER_BG   = "4472C4"   # ヘッダー背景（青）
COLOR_HEADER_FG   = "FFFFFF"   # ヘッダー文字（白）
COLOR_WEEKDAY_ROW = "D9E1F2"   # 曜日行背景

COLOR_REST        = "FFB6C1"   # 休 → ピンク
COLOR_YELLOW      = "FFFF99"   # 認知症加算（黄）
COLOR_PINK_NURSE  = "FFD9E8"   # 中重度加算（薄ピンク）
COLOR_RED_ALERT   = "FF0000"   # 人員不足（赤）

COLOR_SUMMARY_BG  = "F2F2F2"   # 集計行背景

THICK = Side(border_style="thick",  color="000000")
THIN  = Side(border_style="thin",   color="CCCCCC")


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _no_fill() -> PatternFill:
    return PatternFill(fill_type=None)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _border(is_sunday_col: bool) -> Border:
    """日曜列なら右辺をThickにしたBorderを返す。それ以外はThinの格子。"""
    right = THICK if is_sunday_col else THIN
    return Border(left=THIN, right=right, top=THIN, bottom=THIN)


def apply_design(
    wb: Workbook,
    ws,
    schedule_data: dict,
    staff_rows: dict,
    summary_start_row: int,
    summary_count: int,
    validation_result,
):
    year   = schedule_data["year"]
    month  = schedule_data["month"]
    dates: list[datetime.date] = schedule_data["dates"]
    staff_list = schedule_data["staff_list"]
    schedule   = schedule_data["schedule"]

    vr          = validation_result
    red_cells   = vr.red_cells   if vr else set()
    yellow_cells = vr.yellow_cells if vr else set()
    pink_cells  = vr.pink_cells  if vr else set()

    NAME_COL      = 1
    DATE_START_COL = 2
    HEADER_ROW    = 1
    WEEKDAY_ROW   = 2
    total_rows    = summary_start_row + summary_count - 1
    total_cols    = DATE_START_COL + len(dates) - 1

    # 日曜日の列セット（境界判定用）
    sunday_dates = {d for d in dates if d.weekday() == 6}
    # 列番号→日付の逆引き（DATE_START_COL以降）
    col_to_date = {DATE_START_COL + i: d for i, d in enumerate(dates)}

    def is_sun_col(col: int) -> bool:
        d = col_to_date.get(col)
        return d in sunday_dates if d else False

    # ── 列幅・行高 ──────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(NAME_COL)].width = 12
    for i in range(len(dates)):
        ws.column_dimensions[get_column_letter(DATE_START_COL + i)].width = 4.5
    ws.row_dimensions[HEADER_ROW].height = 18
    ws.row_dimensions[WEEKDAY_ROW].height = 16
    for row in staff_rows.values():
        ws.row_dimensions[row].height = 16

    # ── ヘッダー行（日付番号）────────────────────────────────────────
    for col in range(NAME_COL, total_cols + 1):
        cell = ws.cell(row=HEADER_ROW, column=col)
        cell.fill      = _fill(COLOR_HEADER_BG)
        cell.font      = _font(bold=True, color=COLOR_HEADER_FG, size=10)
        cell.alignment = _center()
        cell.border    = _border(is_sun_col(col))

    # ── 曜日行 ────────────────────────────────────────────────────────
    ws.cell(row=WEEKDAY_ROW, column=NAME_COL).border = _border(False)
    for i, d in enumerate(dates):
        col  = DATE_START_COL + i
        cell = ws.cell(row=WEEKDAY_ROW, column=col)
        cell.alignment = _center()
        cell.border    = _border(is_sun_col(col))
        if d.weekday() == 6:      # 日曜
            cell.fill = _fill(COLOR_WEEKDAY_ROW)
            cell.font = _font(color="FF0000", size=9, bold=True)
        elif d.weekday() == 5:    # 土曜
            cell.fill = _fill(COLOR_WEEKDAY_ROW)
            cell.font = _font(color="0000FF", size=9)
        else:
            cell.fill = _fill(COLOR_WEEKDAY_ROW)
            cell.font = _font(size=9)

    # ── 名前列（職種ごとに薄い背景）─────────────────────────────────
    role_colors = {
        "管理者": "D9D9D9", "相談員": "DEEAF1",
        "介護":   "FFFFFF", "看護":   "E2EFDA",
        "OT":     "FFF2CC", "送迎":   "FFF2CC",
        "皿洗い": "FDE9D9", "事務":   "D9D9D9",
    }
    for s in staff_list:
        row  = staff_rows[s.name]
        cell = ws.cell(row=row, column=NAME_COL)
        cell.font      = _font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill      = _fill(role_colors.get(s.role, "FFFFFF"))
        cell.border    = _border(False)

    # ── シフトセル ────────────────────────────────────────────────────
    for s in staff_list:
        row = staff_rows[s.name]
        for i, d in enumerate(dates):
            col   = DATE_START_COL + i
            cell  = ws.cell(row=row, column=col)
            shift = schedule[s.name].get(d, "")
            cell.alignment = _center()
            cell.font      = _font(bold=(shift in ("準", "深")), size=9)
            cell.border    = _border(is_sun_col(col))

            # 優先: 人員不足セル（赤）
            if (s.name, d) in red_cells:
                cell.fill = _fill(COLOR_RED_ALERT)
                cell.font = _font(color="FFFFFF", bold=True, size=9)
            # 認知症加算（黄）
            elif (s.name, d) in yellow_cells:
                cell.fill = _fill(COLOR_YELLOW)
            # 中重度加算（薄ピンク）
            elif (s.name, d) in pink_cells:
                cell.fill = _fill(COLOR_PINK_NURSE)
            # 全てのシフト → 背景なし
            else:
                cell.fill = _no_fill()

    # ── 集計行 ────────────────────────────────────────────────────────
    for j in range(summary_count + 1):
        row = summary_start_row - 1 + j
        ws.row_dimensions[row].height = 15
        label_cell = ws.cell(row=row, column=NAME_COL)
        label_cell.fill      = _fill(COLOR_SUMMARY_BG)
        label_cell.font      = _font(bold=True, size=9)
        label_cell.border    = _border(False)
        for i in range(len(dates)):
            col  = DATE_START_COL + i
            cell = ws.cell(row=row, column=col)
            cell.fill      = _fill(COLOR_SUMMARY_BG)
            cell.alignment = _center()
            cell.font      = _font(size=9)
            cell.border    = _border(is_sun_col(col))

    # ── 外枠太線（上下左右）─────────────────────────────────────────
    for col in range(NAME_COL, total_cols + 1):
        # 1行目の上辺
        c = ws.cell(row=1, column=col)
        b = c.border
        c.border = Border(left=b.left, right=b.right, top=THICK, bottom=b.bottom)
        # 最終行の下辺
        c = ws.cell(row=total_rows, column=col)
        b = c.border
        c.border = Border(left=b.left, right=b.right, top=b.top, bottom=THICK)

    for row in range(1, total_rows + 1):
        # 名前列の左辺
        c = ws.cell(row=row, column=NAME_COL)
        b = c.border
        c.border = Border(left=THICK, right=b.right, top=b.top, bottom=b.bottom)
        # 最終列の右辺
        c = ws.cell(row=row, column=total_cols)
        b = c.border
        c.border = Border(left=b.left, right=THICK, top=b.top, bottom=b.bottom)

    # ── シート表示設定 ────────────────────────────────────────────────
    ws.freeze_panes = "B3"
    ws.sheet_view.showGridLines = False

    # ── 凡例 ─────────────────────────────────────────────────────────
    legend_row = total_rows + 2
    ws.cell(row=legend_row, column=NAME_COL, value="【凡例】").font = _font(bold=True, size=9)
    legend_items = [
        ("黄", COLOR_YELLOW,     "認知症加算対象者が日勤"),
        ("薄ピンク", COLOR_PINK_NURSE, "中重度加算条件達成（看護師）"),
        ("赤", COLOR_RED_ALERT,  "人員不足（警告）"),
    ]
    for k, (label, color, desc) in enumerate(legend_items):
        r = legend_row + 1 + k
        ws.cell(row=r, column=NAME_COL, value=f"　{desc}").font = _font(size=9)
        mark = ws.cell(row=r, column=NAME_COL + 1, value=label)
        mark.fill      = _fill(color)
        mark.alignment = _center()
        mark.font      = _font(size=9)


if __name__ == "__main__":
    from agent2_scheduler import run_scheduler
    from agent3_validator import validate
    from agent4_exporter import export_to_excel

    data = run_scheduler(2026, 4)
    vr   = validate(data)
    path, wb, ws, staff_rows, sum_row, sum_count = export_to_excel(data, vr)
    apply_design(wb, ws, data, staff_rows, sum_row, sum_count, vr)
    wb.save(path)
    print(f"デザイン適用完了: {path}")
