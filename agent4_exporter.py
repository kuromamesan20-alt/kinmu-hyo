"""
agent4_exporter.py: Excel出力
スケジュールデータをExcelに書き込む（スタイルなし）
"""
import datetime
import calendar
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from demo_display import DEMO_MODE, display_names

OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# 行レイアウト
HEADER_ROW = 1     # 列ヘッダー（日付・曜日）
STAFF_START_ROW = 2  # スタッフ行開始
ROLE_COL = 1       # 職種列（A列）
NAME_COL = 2       # 名前列（B列）
DATE_START_COL = 3  # 日付開始列（C列）

WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]

ROLE_DISPLAY = {
    "管理者": "管理者",
    "相談員": "生活相談員",
    "介護": "介護職員",
    "看護": "看護師",
    "OT": "作業療法士",
    "送迎": "送迎",
    "皿洗い": "お皿洗い",
    "事務": "事務",
}


def export_to_excel(schedule_data: dict, validation_result=None) -> Path:
    year = schedule_data["year"]
    month = schedule_data["month"]
    dates: list[datetime.date] = schedule_data["dates"]
    staff_list = schedule_data["staff_list"]
    schedule: dict = schedule_data["schedule"]
    req_map: dict = schedule_data.get("req_map", {})
    display_name_map: dict[str, str] = schedule_data.get("display_names", {})
    prev_month_deep_staff = schedule_data.get("prev_month_deep_staff")
    if DEMO_MODE and not display_name_map:
        display_name_map = display_names([s.name for s in staff_list])

    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = f"{year}年{month}月"

    # ── ヘッダー行（職種・名前列）─────────────────────────────────────
    ws.cell(row=HEADER_ROW, column=ROLE_COL, value="職種")
    ws.cell(row=HEADER_ROW, column=NAME_COL, value="氏名")
    ws.cell(row=HEADER_ROW + 1, column=ROLE_COL, value="")
    ws.cell(row=HEADER_ROW + 1, column=NAME_COL, value="")

    # ── 日付ヘッダー ─────────────────────────────────────────────────
    for i, d in enumerate(dates):
        col = DATE_START_COL + i
        ws.cell(row=HEADER_ROW, column=col, value=d.day)
        ws.cell(row=HEADER_ROW + 1, column=col, value=WEEKDAY_JP[d.weekday()])

    # ── 個人集計列ヘッダー ────────────────────────────────────────────
    PERSONAL_SUMMARY_LABELS = ["早出", "日勤", "A/P", "準夜", "深夜", "休み"]
    personal_summary_start_col = DATE_START_COL + len(dates)
    for k, label in enumerate(PERSONAL_SUMMARY_LABELS):
        col = personal_summary_start_col + k
        ws.cell(row=HEADER_ROW, column=col, value=label)
        ws.cell(row=HEADER_ROW + 1, column=col, value="")

    # ── スタッフ行 ────────────────────────────────────────────────────
    staff_rows: dict[str, int] = {}
    for row_idx, s in enumerate(staff_list):
        row = STAFF_START_ROW + 1 + row_idx  # +1 for weekday row
        staff_rows[s.name] = row
        ws.cell(row=row, column=ROLE_COL, value=ROLE_DISPLAY.get(s.role, s.role))
        ws.cell(row=row, column=NAME_COL, value=display_name_map.get(s.name, s.name))

        for i, d in enumerate(dates):
            col = DATE_START_COL + i
            shift = schedule[s.name].get(d, "")
            req = req_map.get(s.name, {}).get(d)
            is_prev_month_deep_rest = s.name == prev_month_deep_staff and d == dates[0]
            if shift == "休" and (
                is_prev_month_deep_rest or not (req and req.req_type == "希望休")
            ):
                value = None
            elif shift == "皿洗い":
                value = "○"
            else:
                value = shift
            ws.cell(row=row, column=col, value=value)

        # 個人集計（手入力修正に追従するよう数式で集計）
        first_day_col = get_column_letter(DATE_START_COL)
        last_day_col = get_column_letter(DATE_START_COL + len(dates) - 1)
        date_range = f"{first_day_col}{row}:{last_day_col}{row}"
        personal_formulas = [
            f'=COUNTIF({date_range},"早")',
            f'=COUNTIF({date_range},"日")',
            f'=COUNTIF({date_range},"A")+COUNTIF({date_range},"P")',
            f'=COUNTIF({date_range},"準")',
            f'=COUNTIF({date_range},"深")',
            f'=COUNTBLANK({date_range})+COUNTIF({date_range},"休")',
        ]
        for k, formula in enumerate(personal_formulas):
            ws.cell(row=row, column=personal_summary_start_col + k, value=formula)

    # ── 集計行 ────────────────────────────────────────────────────────
    summary_start_row = STAFF_START_ROW + 1 + len(staff_list) + 1
    summary_labels = ["早", "午前(日+A)", "午後(日+P)", "準", "深", "夕・送迎"]
    summary_keys = ["早", "am", "pm", "準", "深", "夕送迎"]

    ws.cell(row=summary_start_row - 1, column=ROLE_COL, value="")
    ws.cell(row=summary_start_row - 1, column=NAME_COL, value="──集計──")

    for j, (label, key) in enumerate(zip(summary_labels, summary_keys)):
        sum_row = summary_start_row + j
        ws.cell(row=sum_row, column=ROLE_COL, value="")
        ws.cell(row=sum_row, column=NAME_COL, value=label)

        for i, d in enumerate(dates):
            col = DATE_START_COL + i
            formula = _summary_formula(key, col, staff_list, staff_rows)
            ws.cell(row=sum_row, column=col, value=formula)

    # ── メタデータ（行・列マップをシートに保存） ──────────────────────
    wb.custom_doc_props = {}  # placeholder

    output_path = OUTPUT_DIR / f"勤務表_{year}年{month}月.xlsx"
    wb.save(output_path)
    print(f"[export] 保存: {output_path}")
    return output_path, wb, ws, staff_rows, summary_start_row, len(summary_labels)


def _summary_formula(key: str, col: int, staff_list: list, staff_rows: dict[str, int]) -> str:
    """集計行の数式"""
    col_letter = get_column_letter(col)
    first_row = min(staff_rows.values())
    last_row = max(staff_rows.values())
    full_range = f"{col_letter}{first_row}:{col_letter}{last_row}"

    def countif_rows(names: list[str], shift: str) -> str:
        if not names:
            return "0"
        return "+".join(
            f'COUNTIF({col_letter}{staff_rows[name]},"{shift}")'
            for name in names
        )

    if key == "早":
        return f'=COUNTIF({full_range},"早")'
    if key == "am":
        names = [s.name for s in staff_list if not s.count_excluded]
        return f'={countif_rows(names, "日")}+{countif_rows(names, "A")}'
    if key == "pm":
        names = [s.name for s in staff_list if not s.count_excluded]
        return f'={countif_rows(names, "日")}+{countif_rows(names, "P")}'
    if key == "準":
        return f'=COUNTIF({full_range},"準")'
    if key == "深":
        return f'=COUNTIF({full_range},"深")'
    if key == "夕送迎":
        return f'=COUNTIF({full_range},"夕")+COUNTIF({full_range},"送迎")+COUNTIF({full_range},"朝夕")'
    return "=0"


if __name__ == "__main__":
    from agent2_scheduler import run_scheduler
    data = run_scheduler(2026, 4)
    result = export_to_excel(data)
    print("Export done:", result[0])
